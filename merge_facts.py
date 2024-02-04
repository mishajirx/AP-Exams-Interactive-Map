from json import dumps, loads
from pprint import pprint
from openpyxl import load_workbook, Workbook

main_wb = Workbook()
main_sheet = main_wb.active
main_sheet.append(["ID",
                   "Subject_ID",
                   "Hierarchy_Level",
                   "School Code",
                   "Tests Taken",
                   "Score=1",
                   "Score=2",
                   "Score=3",
                   "Score=4",
                   "Score=5",
                   "% Score 1-2",
                   "% Score 3-5"])

for subject_ind in range(50):
    print(subject_ind)
    input_file = f"FactsUpd/ap_performance ({subject_ind}).xlsx"

    # load excel file
    workbook = load_workbook(filename=input_file)

    # open workbook
    sheet = workbook.active
    # help(sheet)

    is_first = True
    for row in sheet.values:
        if not is_first:
            main_sheet.append(row)
        is_first = False


main_wb.save("ap_performance_facts.xlsx")
