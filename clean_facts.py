from json import dumps, loads
from pprint import pprint
from openpyxl import load_workbook

common_id = 0

for subject_ind in range(50):
    print(subject_ind)
    input_file = f"Facts/ap_performance ({subject_ind}).xlsx"
    output_file = f"FactsUpd/ap_performance ({subject_ind}).xlsx"

    # load excel file
    workbook = load_workbook(filename=input_file)

    # open workbook
    sheet = workbook.active
    # help(sheet)

    sheet.unmerge_cells(range_string="A1:J1")
    sheet.delete_rows(1)
    sheet.delete_cols(1)

    n = sheet.max_row + 1

    # Replace all percentage periods with commas
    for i in range(2, n):
        sheet[f'H{i}'] = ",".join(str(sheet[f'H{i}'].value).split('.'))
        sheet[f'I{i}'] = ",".join(str(sheet[f'I{i}'].value).split('.'))

    # Remove commas in number of test takers

    for i in range(2, n):
        sheet[f'B{i}'] = "".join(str(sheet[f'B{i}'].value).split(','))

    sheet.insert_cols(1)
    sheet["A1"] = "ID"
    sheet.insert_cols(2)
    sheet["B1"] = "Subject_ID"
    sheet.insert_cols(3)
    sheet["C1"] = "Hierarchy_Level"
    for i in range(2, n):
        sheet[f"A{i}"] = common_id + 1
        common_id += 1
        sheet[f"B{i}"] = subject_ind
        if subject_ind == 0:
            sheet[f"C{i}"] = 0
        elif subject_ind in [1, 7, 10, 20, 31, 38, 47]:
            sheet[f"C{i}"] = 1
        else:
            sheet[f"C{i}"] = 2

    # save the file
    workbook.save(filename=output_file)
