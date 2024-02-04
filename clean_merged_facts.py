from json import dumps, loads
from pprint import pprint
from openpyxl import load_workbook

common_id = 0

input_file = f"ap_performance_facts_poor.xlsx"
output_file = f"ap_performance_facts_poor_new.xlsx"

wb = load_workbook(filename=input_file)
sh = wb.active

for row in range(2, sh.max_row + 1):
    for col in ["D", "E", "F", "G", "H", "I", "J"]:
        if sh[col + str(row)].value:
            sh[col + str(row)] = int(sh[col + str(row)].value.replace(',', ''))
    for col in ["K", "L"]:
        if sh[col + str(row)].value:
            sh[col + str(row)] = float(sh[col + str(row)].value.replace(',', '.'))

wb.save(filename=output_file)